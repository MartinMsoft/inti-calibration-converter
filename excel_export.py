import io
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def has_decimals(vols: dict) -> bool:
    sample = list(vols.values())[:30]
    return any(isinstance(v, float) and v != int(v) for v in sample)


def sanitize_filename_part(text: str) -> str:
    """Limpia caracteres invalidos para nombres de archivo en Windows/Excel."""
    cleaned = re.sub(r'[\\/*?:"<>|]', "_", text).strip()
    return cleaned or "SIN_NOMBRE"


def generate_excel(vols: dict, tank_name: str, cert_number: str,
                    validation: dict, passes_info: str = "",
                    scale_fixes: Optional[list[str]] = None,
                    log_lines: Optional[list[str]] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"TK-{tank_name}"[:31]

    TITLE_FILL = PatternFill("solid", start_color="0D2A4A")
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    ERROR_FILL = PatternFill("solid", start_color="FF4444")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:B1")
    ws["A1"] = f"TABLA DE LLENADO - TANQUE {tank_name}  |  ANTIVARI"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    estado = "APROBADA" if validation["ok"] else "CON ERRORES - ver hoja VALIDACION"
    info = [
        ("Razon Social:", "Antivari S.A."),
        ("Tanque N:", tank_name),
        ("Certificado INTI N:", cert_number),
        ("Unidad:", "dm3"),
        ("Validacion:", estado),
    ]
    for i, (lbl, val) in enumerate(info):
        r = i + 2
        ws.cell(r, 1, lbl).font = Font(name="Arial", bold=True, size=9)
        c2 = ws.cell(r, 2, val)
        c2.font = Font(name="Arial", size=9,
                        color="FF0000" if "ERRORES" in str(val) else "000000")
        ws.row_dimensions[r].height = 14

    header_row = len(info) + 2
    for c, h in [(1, "mm"), (2, "dm3")]:
        cell = ws.cell(header_row, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL; cell.alignment = center; cell.border = border
    ws.row_dimensions[header_row].height = 18
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = f"A{header_row + 1}"

    num_format = "#,##0.000" if has_decimals(vols) else "#,##0"
    bad_mm = validation.get("bad_mm", set())
    MISSING_FILL = PatternFill("solid", start_color="FF8C00")

    row = header_row + 1
    for mm in range(0, max(vols.keys()) + 1):
        vol = vols.get(mm)
        if vol is None:
            c1 = ws.cell(row, 1, mm)
            c1.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c1.alignment = center; c1.border = border; c1.fill = MISSING_FILL
            c2 = ws.cell(row, 2, "FALTANTE")
            c2.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c2.alignment = center; c2.border = border; c2.fill = MISSING_FILL
            row += 1; continue

        is_bad = mm in bad_mm
        is_alt = (mm // 10) % 2 == 1
        fill = ERROR_FILL if is_bad else (ALT_FILL if is_alt else None)
        font_color = "FFFFFF" if is_bad else "000000"

        c1 = ws.cell(row, 1, mm)
        c1.font = Font(name="Arial", size=9, bold=is_bad, color=font_color)
        c1.alignment = center; c1.border = border
        if fill: c1.fill = fill

        c2 = ws.cell(row, 2, vol)
        c2.font = Font(name="Arial", size=9, bold=is_bad, color=font_color)
        c2.alignment = center; c2.border = border; c2.number_format = num_format
        if fill: c2.fill = fill

        row += 1

    wv = wb.create_sheet("VALIDACION")
    wv.column_dimensions["A"].width = 22
    wv.column_dimensions["B"].width = 90

    def vrow(r, lbl, val, bold=False, color="000000"):
        wv.cell(r, 1, lbl).font = Font(name="Arial", bold=bold, size=9)
        wv.cell(r, 2, val).font = Font(name="Arial", bold=bold, size=9, color=color)

    r = 1
    vrow(r, "REPORTE DE VALIDACION", f"Tanque {tank_name}", bold=True); r += 1
    vrow(r, "Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")); r += 1
    vrow(r, "Certificado", cert_number); r += 1
    vrow(r, "Procesamiento", passes_info); r += 1
    vrow(r, "Total mm", str(validation["stats"].get("total_mm", "-"))); r += 1
    vrow(r, "Rango", validation["stats"].get("rango", "-")); r += 1
    vrow(r, "Volumen min", validation["stats"].get("vol_min", "-") + " dm3"); r += 1
    vrow(r, "Volumen max", validation["stats"].get("vol_max", "-") + " dm3"); r += 1
    vrow(r, "MM faltantes", str(validation["stats"].get("faltantes", "-"))); r += 1
    r += 1
    result_txt = "APROBADA" if validation["ok"] else "FALLIDA - REVISAR FILAS EN ROJO"
    result_col = "008000" if validation["ok"] else "FF0000"
    vrow(r, "RESULTADO", result_txt, bold=True, color=result_col); r += 2

    if validation["errors"]:
        vrow(r, "ERRORES", "", bold=True, color="FF0000"); r += 1
        for e in validation["errors"]:
            vrow(r, "", e, color="FF0000"); r += 1
        r += 1
    if validation["warnings"]:
        vrow(r, "ADVERTENCIAS", "", bold=True, color="CC6600"); r += 1
        for w in validation["warnings"]:
            vrow(r, "", w, color="CC6600"); r += 1
        r += 1
    if scale_fixes:
        vrow(r, "CORRECCIONES ESCALA", f"{len(scale_fixes)} valor(es)", bold=True, color="8B008B"); r += 1
        for f_txt in scale_fixes[:20]:
            vrow(r, "", f_txt, color="8B008B"); r += 1
        if len(scale_fixes) > 20:
            vrow(r, "", f"... y {len(scale_fixes)-20} mas", color="8B008B"); r += 1
        r += 1
    if log_lines:
        vrow(r, "LOG DE PROCESAMIENTO", f"{len(log_lines)} evento(s)", bold=True, color="555555"); r += 1
        for line in log_lines[:100]:
            vrow(r, "", line, color="555555"); r += 1
        if len(log_lines) > 100:
            vrow(r, "", f"... y {len(log_lines)-100} mas", color="555555"); r += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()
