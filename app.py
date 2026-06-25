import streamlit as st
import anthropic
import json
import io
import base64
import re
from datetime import datetime
from pdf2image import convert_from_bytes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MODEL_FAST    = "claude-haiku-4-5-20251001"
MODEL_PRECISE = "claude-sonnet-4-6"
BATCH_SIZE    = 3   # paginas por llamada (contexto cruzado sin exceder max_tokens)
MAX_TOKENS    = 8192  # suficiente para ~3 paginas de tabla

# ── Helpers de imagen ────────────────────────────────────────────────────────

def pdf_to_images(pdf_bytes: bytes):
    return convert_from_bytes(pdf_bytes, dpi=300)

def image_to_base64(image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")

# ── Prompt ───────────────────────────────────────────────────────────────────

PROMPT = """Las imagenes adjuntas son paginas CONSECUTIVAS de una tabla de calibracion de tanque industrial certificada por INTI (Argentina).

La tabla tiene este formato:
- Primera columna: valor base en mm (0, 10, 20, 30, ...)
- Columnas 0 a 9: los 10 mm individuales de esa fila (base+0 a base+9)
- Valores en dm3

Extrae TODOS los datos de TODAS las paginas en orden, en UN SOLO JSON:
{
  "rows": [
    {"base_mm": 0, "values": [v0, v1, v2, v3, v4, v5, v6, v7, v8, v9]},
    {"base_mm": 10, "values": [v0, v1, v2, v3, v4, v5, v6, v7, v8, v9]}
  ]
}

Reglas CRITICAS:

NUMEROS - el punto (.) es separador de MILES, NO decimal:
  788.068   -> entero 788068
  1.160.362 -> entero 1160362
  Si hay DOS o mas puntos, siempre es entero: quita los puntos.
  Si hay UN solo punto y el numero es menor a 10000, es decimal: 27.344

CONSISTENCIA entre paginas:
  Los volumenes siempre aumentan de fila en fila. Nunca disminuyen.
  Si el ultimo valor de una pagina es 137914, la siguiente pagina debe continuar
  en la misma escala: 138016, 138118, etc. (NO 138.016 ni 138016000).

IGNORAR completamente:
  Numeros de pagina, encabezados, pies de pagina, firmas, sellos, logos.
  Todo texto fuera de la tabla de datos.

FORMATO:
  Cada fila tiene EXACTAMENTE 10 valores (null si la celda esta vacia).
  No incluyas filas de paginas sin tabla (caratula, texto libre, firma).
  Responde UNICAMENTE con el JSON, sin texto adicional ni bloques de codigo."""


def make_retry_prompt(expected_start: int, expected_end: int) -> str:
    return PROMPT + f"""

ATENCION ESPECIAL PARA ESTE BLOQUE:
Se esperan datos en el rango aproximado mm={expected_start} a mm={expected_end}.
Revisa cada digito con maxima precision. Los volumenes deben ser enteros grandes
(ej: 306592, 306694...), NO decimales (ej: 306.592). Es una calibracion industrial critica."""

# ── Extraccion por lotes ─────────────────────────────────────────────────────

def extract_batch(client: anthropic.Anthropic, batch: list,
                  model: str = MODEL_FAST, prompt: str = PROMPT) -> list[dict]:
    """
    Extrae datos de un lote de paginas consecutivas en UNA sola llamada API.
    batch = lista de (page_num, image)
    Retorna lista de rows (todos los de todas las paginas del lote en orden).
    """
    content = []
    for page_num, img in batch:
        content.append({"type": "text", "text": f"[Pagina {page_num}]"})
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png",
                       "data": image_to_base64(img)}
        })
    content.append({"type": "text", "text": prompt})

    page_nums = [p for p, _ in batch]
    label = f"{page_nums[0]}-{page_nums[-1]}"
    try:
        response = client.messages.create(
            model=model,
            max_tokens=MAX_TOKENS,
            messages=[{"role": "user", "content": content}],
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        # Si Claude agrego texto antes del JSON, buscar el inicio del objeto
        if not raw.startswith("{"):
            idx = raw.find("{")
            if idx != -1:
                raw = raw[idx:]
        data = json.loads(raw)
        return data.get("rows", [])
    except Exception as e:
        st.warning(f"Paginas {label}: error ({e}). Se omite.")
        return []

# ── Construccion del diccionario mm->dm3 ─────────────────────────────────────

def build_vols(all_rows: list[dict]) -> dict[int, float]:
    """Convierte la lista plana de rows en el diccionario mm->volumen."""
    vols = {}
    for row in all_rows:
        base = int(row["base_mm"])
        for i, v in enumerate(row["values"][:10]):
            if v is not None:
                vols[base + i] = v
    return vols

# ── Correccion de escala ──────────────────────────────────────────────────────

def fix_scale_errors(vols: dict) -> tuple[dict[int, float], list[str]]:
    """
    Corrige errores x1000 (punto de miles leido como decimal) usando monotonia.
    Pasada inversa: si vol[mm] > vol[mm+1]*500 -> divide por 1000.
    Pasada directa: si vol[mm] < vol[mm-1] y vol[mm]*1000 restaura monotonia -> multiplica.
    """
    mm_sorted = sorted(vols.keys())
    if len(mm_sorted) < 2:
        return vols, []

    corrected = dict(vols)
    fixes = []

    for i in range(len(mm_sorted) - 2, -1, -1):
        mm      = mm_sorted[i]
        mm_next = mm_sorted[i + 1]
        v       = corrected[mm]
        v_next  = corrected[mm_next]
        if v > v_next * 500:
            v_down = round(v / 1000, 3)
            if v_down <= v_next and v_down >= v_next / 2:
                corrected[mm] = v_down
                fixes.append(f"mm={mm}: {v} -> {v_down} (div1000)")

    for i in range(1, len(mm_sorted)):
        mm      = mm_sorted[i]
        mm_prev = mm_sorted[i - 1]
        v       = corrected[mm]
        v_prev  = corrected[mm_prev]
        if v < v_prev:
            v_up = v * 1000
            if v_up >= v_prev and v_up <= v_prev * 2:
                corrected[mm] = v_up
                fixes.append(f"mm={mm}: {v} -> {v_up} (x1000)")

    return corrected, fixes

# ── Validacion ────────────────────────────────────────────────────────────────

def validate_vols(vols: dict) -> dict:
    errors, warnings = [], []
    if not vols:
        return {"ok": False, "errors": ["No hay datos."], "warnings": [],
                "stats": {}, "bad_mm": set(), "missing": []}

    mm_sorted = sorted(vols.keys())
    min_mm, max_mm = mm_sorted[0], mm_sorted[-1]

    # 1. MM faltantes
    missing = [mm for mm in range(min_mm, max_mm + 1) if mm not in vols]
    if missing:
        groups, start, prev = [], missing[0], missing[0]
        for m in missing[1:]:
            if m == prev + 1:
                prev = m
            else:
                groups.append((start, prev)); start = prev = m
        groups.append((start, prev))
        ranges_str = ", ".join(str(a) if a == b else f"{a}-{b}" for a, b in groups[:10])
        if len(groups) > 10:
            ranges_str += f" ... y {len(groups)-10} rangos mas"
        errors.append(f"MM faltantes ({len(missing)} valores): {ranges_str}")

    # 2. Volumen siempre creciente
    non_mono = []
    prev_vol = None
    for mm in mm_sorted:
        vol = vols[mm]
        if prev_vol is not None and vol < prev_vol:
            non_mono.append((mm, prev_vol, vol))
        prev_vol = vol
    if non_mono:
        detail = "; ".join(f"mm={mm}: {pv:.3f}->{v:.3f}" for mm, pv, v in non_mono[:5])
        if len(non_mono) > 5: detail += f" ... y {len(non_mono)-5} mas"
        errors.append(f"Volumen decrece en {len(non_mono)} punto(s): {detail}")

    # 3. Incrementos anomalos (mediana local, ventana +-30 mm)
    inc_map: dict[int, float] = {}
    for i in range(1, len(mm_sorted)):
        if mm_sorted[i] == mm_sorted[i - 1] + 1:
            inc_map[mm_sorted[i]] = vols[mm_sorted[i]] - vols[mm_sorted[i - 1]]

    outliers = []
    if inc_map:
        inc_keys = sorted(inc_map.keys())
        WINDOW    = 30
        THRESHOLD = 4.0

        def median(lst):
            s = sorted(lst)
            return s[len(s) // 2]

        for idx, mm in enumerate(inc_keys):
            actual = inc_map[mm]
            neighbors = [
                inc_map[inc_keys[j]]
                for j in range(max(0, idx - WINDOW), min(len(inc_keys), idx + WINDOW + 1))
                if j != idx and inc_map[inc_keys[j]] > 0
            ]
            if len(neighbors) < 5:
                continue
            local_med = median(neighbors)
            if local_med <= 0:
                continue
            ratio = actual / local_med
            if ratio > THRESHOLD or ratio < 1 / THRESHOLD:
                outliers.append((mm, actual, local_med, ratio))

        if outliers:
            detail = "; ".join(
                f"mm={mm}: D={act:.3f} (esp~{exp:.3f}, ratio={rat:.1f}x)"
                for mm, act, exp, rat in outliers[:8]
            )
            if len(outliers) > 8:
                detail += f" ... y {len(outliers) - 8} mas"
            errors.append(f"Incremento anomalo en {len(outliers)} punto(s): {detail}")

    bad_mm = set(missing)
    bad_mm.update(mm for mm, _, _ in non_mono)
    bad_mm.update(mm for mm, _, _, _ in outliers)

    stats = {
        "total_mm": len(vols),
        "rango": f"{min_mm} - {max_mm} mm",
        "vol_min": f"{min(vols.values()):.3f}",
        "vol_max": f"{max(vols.values()):.3f}",
        "faltantes": len(missing),
    }
    return {"ok": len(errors) == 0, "errors": errors, "warnings": warnings,
            "stats": stats, "bad_mm": bad_mm, "missing": missing}

# ── Logica de retry ───────────────────────────────────────────────────────────

def find_batches_to_retry(batch_results: list, validation: dict) -> set[int]:
    """
    Retorna indices (en batch_results) de lotes que deben re-procesarse con Sonnet.
    batch_results = [(batch_idx, [(page_num, img)], rows), ...]
    """
    bad_mm    = validation["bad_mm"]
    missing_s = set(validation["missing"])
    retry_idx = set()

    for b_idx, batch, rows in batch_results:
        # Lotes sin ninguna fila
        if not rows:
            retry_idx.add(b_idx)
            continue
        # Lotes con mm problematicos
        for row in rows:
            base = int(row["base_mm"])
            if any((base + i) in bad_mm for i in range(10)):
                retry_idx.add(b_idx)
                break

    # Lotes adyacentes a gaps (interseccion con el rango faltante +-20mm)
    if missing_s:
        min_m, max_m = min(missing_s), max(missing_s)
        for b_idx, batch, rows in batch_results:
            if not rows: continue
            bases = [int(r["base_mm"]) for r in rows]
            batch_min = min(bases)
            batch_max = max(bases) + 9
            if batch_max >= min_m - 20 and batch_min <= max_m + 20:
                retry_idx.add(b_idx)

    return retry_idx


def batch_mm_range(rows: list) -> tuple[int, int]:
    """Rango mm cubierto por un lote de rows."""
    if not rows:
        return 0, 0
    bases = [int(r["base_mm"]) for r in rows]
    return min(bases), max(bases) + 9

# ── Generacion del Excel ──────────────────────────────────────────────────────

def has_decimals(vols: dict) -> bool:
    sample = list(vols.values())[:30]
    return any(isinstance(v, float) and v != int(v) for v in sample)

def generate_excel(vols: dict, tank_name: str, cert_number: str,
                   validation: dict, passes_info: str = "",
                   scale_fixes: list[str] = None) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = f"TK-{tank_name}"

    TITLE_FILL  = PatternFill("solid", start_color="0D2A4A")
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
    ERROR_FILL  = PatternFill("solid", start_color="FF4444")
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:B1")
    ws["A1"] = f"TABLA DE LLENADO - TANQUE {tank_name}  |  TAGSA"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    estado = "APROBADA" if validation["ok"] else "CON ERRORES - ver hoja VALIDACION"
    info = [
        ("Razon Social:", "Antivari S.A."),
        ("Tanque N:", tank_name),
        ("Certificado INTI N:", cert_number),
        ("Unidad:", "dm3"),
        ("Validacion:", estado),
    ]
    for i, (lbl, val) in enumerate(info):
        r = i + 2
        ws.cell(r, 1, lbl).font = Font(name="Arial", bold=True, size=9)
        c2 = ws.cell(r, 2, val)
        c2.font = Font(name="Arial", size=9,
                       color="FF0000" if "ERRORES" in str(val) else "000000")
        ws.row_dimensions[r].height = 14

    header_row = len(info) + 2
    for c, h in [(1, "mm"), (2, "dm3")]:
        cell = ws.cell(header_row, c, h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[header_row].height = 18
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = f"A{header_row + 1}"

    num_format   = "#,##0.000" if has_decimals(vols) else "#,##0"
    bad_mm       = validation.get("bad_mm", set())
    MISSING_FILL = PatternFill("solid", start_color="FF8C00")

    row = header_row + 1
    for mm in range(0, max(vols.keys()) + 1):
        vol = vols.get(mm)

        if vol is None:
            c1 = ws.cell(row, 1, mm)
            c1.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c1.alignment = center; c1.border = border; c1.fill = MISSING_FILL
            c2 = ws.cell(row, 2, "FALTANTE")
            c2.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c2.alignment = center; c2.border = border; c2.fill = MISSING_FILL
            row += 1
            continue

        is_bad = mm in bad_mm
        is_alt = (mm // 10) % 2 == 1
        fill = ERROR_FILL if is_bad else (ALT_FILL if is_alt else None)
        font_color = "FFFFFF" if is_bad else "000000"

        c1 = ws.cell(row, 1, mm)
        c1.font = Font(name="Arial", size=9, bold=is_bad, color=font_color)
        c1.alignment = center; c1.border = border
        if fill: c1.fill = fill

        c2 = ws.cell(row, 2, vol)
        c2.font = Font(name="Arial", size=9, bold=is_bad, color=font_color)
        c2.alignment = center; c2.border = border
        c2.number_format = num_format
        if fill: c2.fill = fill

        row += 1

    # ── Hoja VALIDACION ──────────────────────────────────────────────────────
    wv = wb.create_sheet("VALIDACION")
    wv.column_dimensions["A"].width = 22
    wv.column_dimensions["B"].width = 90

    def vrow(r, lbl, val, bold=False, color="000000"):
        wv.cell(r, 1, lbl).font = Font(name="Arial", bold=bold, size=9)
        wv.cell(r, 2, val).font = Font(name="Arial", bold=bold, size=9, color=color)

    r = 1
    vrow(r, "REPORTE DE VALIDACION", f"Tanque {tank_name}", bold=True); r += 1
    vrow(r, "Fecha",        datetime.now().strftime("%Y-%m-%d %H:%M")); r += 1
    vrow(r, "Certificado",  cert_number); r += 1
    vrow(r, "Procesamiento", passes_info); r += 1
    vrow(r, "Total mm",     str(validation["stats"].get("total_mm", "-"))); r += 1
    vrow(r, "Rango",        validation["stats"].get("rango", "-")); r += 1
    vrow(r, "Volumen min",  validation["stats"].get("vol_min", "-") + " dm3"); r += 1
    vrow(r, "Volumen max",  validation["stats"].get("vol_max", "-") + " dm3"); r += 1
    vrow(r, "MM faltantes", str(validation["stats"].get("faltantes", "-"))); r += 1
    r += 1

    result_txt = "APROBADA" if validation["ok"] else "FALLIDA - REVISAR FILAS EN ROJO"
    result_col = "008000" if validation["ok"] else "FF0000"
    vrow(r, "RESULTADO", result_txt, bold=True, color=result_col); r += 2

    if validation["errors"]:
        vrow(r, "ERRORES", "", bold=True, color="FF0000"); r += 1
        for e in validation["errors"]:
            vrow(r, "", e, color="FF0000"); r += 1
        r += 1

    if validation["warnings"]:
        vrow(r, "ADVERTENCIAS", "", bold=True, color="CC6600"); r += 1
        for w in validation["warnings"]:
            vrow(r, "", w, color="CC6600"); r += 1
        r += 1

    if scale_fixes:
        vrow(r, "CORRECCIONES ESCALA", f"{len(scale_fixes)} valor(es)", bold=True, color="8B008B"); r += 1
        for f_txt in scale_fixes[:20]:
            vrow(r, "", f_txt, color="8B008B"); r += 1
        if len(scale_fixes) > 20:
            vrow(r, "", f"... y {len(scale_fixes)-20} mas", color="8B008B"); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── UI ────────────────────────────────────────────────────────────────────────

def get_api_key() -> str:
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        return ""

def main():
    st.set_page_config(
        page_title="Conversor Tablas INTI - TAGSA",
        page_icon="=",
        layout="centered",
    )
    st.title("Conversor Tablas de Calibracion INTI")
    st.caption("Antivari S.A.")
    st.divider()

    secret_key = get_api_key()
    api_key = secret_key or st.text_input(
        "API Key de Anthropic", type="password",
        help="Configurala como secret en Streamlit Cloud para no escribirla cada vez."
    )

    col1, col2 = st.columns(2)
    with col1:
        tank_name   = st.text_input("N de Tanque",        placeholder="ej: TK-81")
    with col2:
        cert_number = st.text_input("N Certificado INTI", placeholder="ej: INTI 2623")

    uploaded = st.file_uploader("Subi el PDF del certificado INTI", type="pdf")
    ready = bool(api_key and tank_name and uploaded)

    if st.button("Convertir a Excel", type="primary", disabled=not ready):
        client = anthropic.Anthropic(api_key=api_key)
        pdf_bytes = uploaded.read()

        # ── PASADA 1: Haiku (lotes de BATCH_SIZE paginas) ───────────────────
        with st.status(f"Pasada 1 - lectura rapida (lotes de {BATCH_SIZE} paginas)...",
                       expanded=True) as status1:
            st.write("Convirtiendo PDF a imagenes...")
            images = pdf_to_images(pdf_bytes)
            total_pages = len(images)
            st.write(f"PDF tiene {total_pages} pagina(s). Procesando en lotes de {BATCH_SIZE}...")

            batch_results = []   # lista de (batch_idx, [(page_num, img)], rows)
            all_rows_p1  = []

            for b_start in range(0, total_pages, BATCH_SIZE):
                b_end    = min(b_start + BATCH_SIZE, total_pages)
                b_idx    = b_start // BATCH_SIZE
                batch    = [(b_start + j + 1, images[b_start + j]) for j in range(b_end - b_start)]
                pnums    = [p for p, _ in batch]
                label    = f"{pnums[0]}-{pnums[-1]}" if len(pnums) > 1 else str(pnums[0])
                st.write(f"Paginas {label} ({len(batch)} imagenes en una sola llamada)...")
                rows = extract_batch(client, batch, model=MODEL_FAST, prompt=PROMPT)
                batch_results.append((b_idx, batch, rows))
                all_rows_p1.extend(rows)
                st.write(f"  -> {len(rows)} filas extraidas.")

            vols           = build_vols(all_rows_p1)
            vols, sf1      = fix_scale_errors(vols)
            if sf1:
                st.info(f"Escala corregida en {len(sf1)} valores.")
            validation_1   = validate_vols(vols)
            p1_ok          = validation_1["ok"]
            p1_lbl         = "Pasada 1 completa - sin errores" if p1_ok else f"Pasada 1: {len(validation_1['errors'])} error(es)"
            status1.update(label=p1_lbl, state="complete")

        # ── PASADA 2: Sonnet (solo lotes con problemas) ──────────────────────
        batches_to_retry = find_batches_to_retry(batch_results, validation_1)
        passes_info = "1 pasada (Haiku, lotes)"

        if batches_to_retry and not p1_ok:
            passes_info = f"2 pasadas - Haiku + Sonnet en {len(batches_to_retry)} lote(s)"
            with st.status(f"Pasada 2 - re-procesando {len(batches_to_retry)} lote(s) con Sonnet...",
                           expanded=True) as status2:

                for b_idx, batch, old_rows in batch_results:
                    if b_idx not in batches_to_retry:
                        continue
                    exp_start, exp_end = batch_mm_range(old_rows)
                    pnums = [p for p, _ in batch]
                    label = f"{pnums[0]}-{pnums[-1]}" if len(pnums) > 1 else str(pnums[0])
                    st.write(f"Re-procesando paginas {label} con Sonnet (mm ~{exp_start}-{exp_end})...")
                    retry_prompt = make_retry_prompt(exp_start, exp_end)
                    new_rows = extract_batch(client, batch, model=MODEL_PRECISE,
                                            prompt=retry_prompt)
                    # Solo reemplazar si Sonnet devolvio datos (nunca descartar datos de Haiku)
                    keep_rows = new_rows if new_rows else old_rows
                    for i, (bi, bt, _) in enumerate(batch_results):
                        if bi == b_idx:
                            batch_results[i] = (bi, bt, keep_rows)
                            break
                    st.write(f"  -> {len(new_rows)} filas de Sonnet (antes Haiku: {len(old_rows)}). Usando: {len(keep_rows)}.")

                all_rows_p2 = []
                for _, _, rows in batch_results:
                    all_rows_p2.extend(rows)

                vols         = build_vols(all_rows_p2)
                vols, sf2    = fix_scale_errors(vols)
                if sf2:
                    st.info(f"Escala corregida en {len(sf2)} valores.")
                validation_2 = validate_vols(vols)
                p2_lbl       = "Pasada 2 completa - sin errores" if validation_2["ok"] else f"Pasada 2: {len(validation_2['errors'])} error(es) restantes"
                status2.update(label=p2_lbl, state="complete")
            validation   = validation_2
            scale_fixes  = sf2
        else:
            validation   = validation_1
            scale_fixes  = sf1

        if not vols:
            st.error("No se pudieron extraer datos del PDF. Revisa que el PDF tenga tablas de calibracion visibles y vuelve a intentar.")
            st.stop()

        st.write("Generando Excel...")
        excel_bytes = generate_excel(vols, tank_name, cert_number or "-",
                                     validation, passes_info, scale_fixes)

        # ── Reporte ──────────────────────────────────────────────────────────
        st.subheader("Reporte de validacion")
        s = validation["stats"]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total mm",  s["total_mm"])
        c2.metric("Rango",     s["rango"])
        c3.metric("Vol. min",  s["vol_min"] + " dm3")
        c4.metric("Vol. max",  s["vol_max"] + " dm3")

        if validation["ok"] and not validation["warnings"]:
            st.success("Validacion APROBADA - todos los controles pasaron.")
        elif validation["ok"]:
            st.warning("Aprobada con advertencias - revisa los detalles.")
        else:
            st.error("Validacion FALLIDA - el Excel tiene errores (filas en rojo). Revisa antes de usar.")

        if validation["errors"]:
            with st.expander("Errores", expanded=True):
                for e in validation["errors"]:
                    st.error(e)
        if validation["warnings"]:
            with st.expander("Advertencias"):
                for w in validation["warnings"]:
                    st.warning(w)

        st.divider()
        fname = f"TK-{tank_name}_Tabla_Llenado.xlsx"
        st.download_button(
            label="Descargar Excel",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary" if validation["ok"] else "secondary",
        )
        if not validation["ok"]:
            st.caption("El archivo tiene errores. Las filas problematicas estan marcadas en ROJO. Verifica contra el PDF original antes de subir al sistema.")

if __name__ == "__main__":
    main()
